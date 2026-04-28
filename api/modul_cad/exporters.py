from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from .pricing import fetch_uf_value


HEADER_FILL = PatternFill("solid", fgColor="FF50637C")
SECTION_FILL = PatternFill("solid", fgColor="FFAEABAB")
TOTAL_FILL = PatternFill("solid", fgColor="FFEFEFEF")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")
THIN_SIDE = Side(style="thin", color="FF000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def format_clp(value: int | float) -> str:
    return f"$ {int(round(value)):,}".replace(",", ".")


def _apply_row_border(ws, row: int, columns: str) -> None:
    for col in columns:
        ws[f"{col}{row}"].border = THIN_BORDER


def _style_table_header(ws, row: int) -> None:
    for col in ("B", "C", "D", "E"):
        cell = ws[f"{col}{row}"]
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.border = THIN_BORDER


def _set_section_label(ws, left_cell: str, right_range: str, label: str, value: str) -> None:
    ws[left_cell] = label
    ws[left_cell].font = Font(color="FFFFFFFF", bold=True)
    ws[left_cell].fill = SECTION_FILL
    ws[left_cell].alignment = Alignment(horizontal="center", wrap_text=True)
    ws[left_cell].border = THIN_BORDER

    ws.merge_cells(right_range)
    anchor = right_range.split(":")[0]
    ws[anchor] = value
    ws[anchor].fill = WHITE_FILL
    ws[anchor].alignment = Alignment(horizontal="center", wrap_text=True)
    ws[anchor].border = THIN_BORDER


def _build_commercial_items(quote: dict[str, Any]) -> list[tuple[str, float]]:
    geometry = quote["geometry"]
    totals = quote["totals"]
    project = quote["project"]
    kitchen = project["kitchen_type"]
    bathroom = project["bathroom_type"]

    return [
        (
            f"Vivienda modular terminada en fabrica ({geometry.gross_area_m2:.2f} m2) "
            f"con cocina {kitchen} y banos {bathroom}",
            totals["materials_subtotal"] + totals["manufacturing_cost"],
        ),
        ("Instalacion en terreno y terminaciones finales", totals["installation_cost"]),
        ("Fundaciones estimadas para configuracion modul", totals["foundation_cost"]),
        ("Transporte de modulos a sitio", totals["transport_cost"]),
        ("Ingenieria, cubicacion y coordinacion tecnica", totals["engineering_cost"]),
        ("Puesta en obra y logistica base", totals["site_setup_cost"]),
        ("Contingencia tecnica", totals["contingency"]),
        ("Overhead operacional", totals["overhead"]),
        ("Margen comercial", totals["profit"]),
    ]


def export_excel(quote: dict[str, Any], plan_path: str | Path, output_path: str | Path, config: dict[str, Any]) -> Path:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Cotizacion"

    ws.column_dimensions["A"].width = 3.25
    ws.column_dimensions["B"].width = 14.75
    ws.column_dimensions["C"].width = 50.38
    ws.column_dimensions["D"].width = 15.13
    ws.column_dimensions["E"].width = 24.0

    project = quote["project"]
    geometry = quote["geometry"]
    totals = quote["totals"]
    quote_defaults = config["quote_defaults"]
    target_date = date.fromisoformat(quote["generated_on"])
    try:
        uf_value = fetch_uf_value(target_date)
        uf_source = "SII"
    except Exception:
        uf_value = 40053.72
        uf_source = "fallback"

    advance_payment = round(totals["final_total"] * quote_defaults["advance_payment_pct"], 2)
    balance_net = round(totals["final_total"] - advance_payment, 2)
    vat_amount = round(balance_net * quote_defaults["vat_pct"], 2)
    balance_with_vat = round(balance_net + vat_amount, 2)
    total_payable = round(advance_payment + balance_with_vat, 2)
    commercial_items = _build_commercial_items(quote)

    if Path(plan_path).exists():
        plan_image = XLImage(str(plan_path))
        plan_image.width = 285
        plan_image.height = 215
        ws.add_image(plan_image, "C2")

    ws.merge_cells("B21:E21")
    ws["B21"] = "COTIZACION PARA:"
    ws["B21"].font = Font(bold=True)
    ws["B21"].alignment = Alignment(horizontal="center")
    ws["B21"].border = THIN_BORDER

    _set_section_label(ws, "B22", "C22:E22", "Nombre", project["client"]["name"].upper())
    _set_section_label(ws, "B23", "C23:E23", "Correo Electronico:", project["client"].get("email", "-"))
    _set_section_label(ws, "B25", "C25:E25", "Proyecto:", project["project_name"].upper())
    _set_section_label(ws, "B26", "C26:E26", "Region Instalacion", project["location"].upper())
    _set_section_label(ws, "B27", "C27:E27", "Superficie Tiny Modul  (M2)", f"{geometry.gross_area_m2:.2f}")
    _set_section_label(ws, "B28", "C28:E28", "Superficie Terraza (M2)", "0.00")
    _set_section_label(ws, "B29", "C29:E29", "Superficie Total (M2)", f"=VALUE(C27)+VALUE(C28)")

    _style_table_header(ws, 31)
    ws["B31"] = "ITEM"
    ws["C31"] = "DESCRIPCION"
    ws["D31"] = "VALOR EN UF"
    ws["E31"] = "PRECIO"

    start_row = 33
    subtotal_row = start_row + len(commercial_items) + 1
    uf_row = subtotal_row + 10
    for offset, (description, amount_clp) in enumerate(commercial_items, start=0):
        row = start_row + offset
        ws[f"B{row}"] = offset + 1
        ws[f"C{row}"] = description
        ws[f"D{row}"] = round(amount_clp / uf_value, 6)
        ws[f"E{row}"] = f"=D{row}*$E${uf_row}"

        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="bottom")
        ws[f"C{row}"].font = Font(name="Arial", size=11)
        ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="bottom")
        ws[f"D{row}"].font = Font(name="Arial", size=11)
        ws[f"D{row}"].number_format = "0.00"
        ws[f"E{row}"].alignment = Alignment(horizontal="right", vertical="bottom")
        ws[f"E{row}"].font = Font(name="Arial", size=11)
        ws[f"E{row}"].fill = WHITE_FILL
        ws[f"E{row}"].number_format = "#,##0"
        _apply_row_border(ws, row, "BCDE")

    ws[f"C{subtotal_row}"] = "Valor total sin incluir IVA ($)"
    ws[f"D{subtotal_row}"] = f"=SUM(D{start_row}:D{subtotal_row - 2})"
    ws[f"E{subtotal_row}"] = f"=SUM(E{start_row}:E{subtotal_row - 2})"
    for col in ("C", "D", "E"):
        ws[f"{col}{subtotal_row}"].fill = TOTAL_FILL
        ws[f"{col}{subtotal_row}"].border = THIN_BORDER
    ws[f"C{subtotal_row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"D{subtotal_row}"].font = Font(name="Arial", size=11, bold=True)
    ws[f"E{subtotal_row}"].font = Font(name="Arial", size=12, bold=True)
    ws[f"D{subtotal_row}"].alignment = Alignment(horizontal="right")
    ws[f"E{subtotal_row}"].alignment = Alignment(horizontal="right")
    ws[f"D{subtotal_row}"].number_format = "0.00"
    ws[f"E{subtotal_row}"].number_format = "#,##0"

    ws[f"D{subtotal_row + 2}"] = "Anticipo Neto"
    ws[f"E{subtotal_row + 2}"] = advance_payment
    ws[f"D{subtotal_row + 3}"] = "Saldo (Neto)"
    ws[f"E{subtotal_row + 3}"] = balance_net
    ws[f"D{subtotal_row + 4}"] = "IVA (del saldo Neto)"
    ws[f"E{subtotal_row + 4}"] = vat_amount
    ws[f"D{subtotal_row + 5}"] = "Monto de Saldo con IVA"
    ws[f"E{subtotal_row + 5}"] = balance_with_vat
    ws[f"D{subtotal_row + 7}"] = f"Monto Total a Pagar por {project['client']['name']}"
    ws[f"E{subtotal_row + 7}"] = total_payable
    ws[f"D{uf_row}"] = f"Valor UF {target_date.strftime('%d-%m-%Y')}"
    ws[f"E{uf_row}"] = uf_value

    for row in (subtotal_row + 2, subtotal_row + 3, subtotal_row + 4, subtotal_row + 5, subtotal_row + 7, uf_row):
        for col in ("D", "E"):
            ws[f"{col}{row}"].fill = TOTAL_FILL
            ws[f"{col}{row}"].border = THIN_BORDER
            ws[f"{col}{row}"].number_format = "#,##0"
        ws[f"D{row}"].font = Font(bold=row in (subtotal_row + 2, subtotal_row + 7))
        ws[f"E{row}"].font = Font(size=12 if row in (subtotal_row + 2, subtotal_row + 7) else 11, bold=row in (subtotal_row + 2, subtotal_row + 7))
        ws[f"D{row}"].alignment = Alignment(horizontal="center" if row == uf_row else None, wrap_text=row == uf_row)
        ws[f"E{row}"].alignment = Alignment(horizontal="center" if row == uf_row else "right")

    ws[f"E{uf_row}"].number_format = "#,##0.00"

    notes_start = uf_row + 5
    ws[f"C{notes_start}"] = "NOTAS"
    ws[f"C{notes_start}"].fill = HEADER_FILL
    ws[f"C{notes_start}"].font = Font(color="FFFFFFFF", bold=True)
    ws[f"C{notes_start}"].alignment = Alignment(horizontal="center")
    for idx, note in enumerate(quote_defaults["notes"], start=1):
        ws[f"C{notes_start + idx}"] = (
            note.replace("hoy", target_date.strftime("%d-%m-%Y"))
            .replace("dia de emision de la cotizacion", f"{target_date.strftime('%d-%m-%Y')} ({uf_source})")
        )
        ws[f"C{notes_start + idx}"].alignment = Alignment(wrap_text=True)

    payment_start = notes_start + 13
    ws[f"C{payment_start}"] = "FORMA DE PAGO"
    ws[f"C{payment_start}"].fill = HEADER_FILL
    ws[f"C{payment_start}"].font = Font(color="FFFFFFFF", bold=True)
    ws[f"C{payment_start}"].alignment = Alignment(horizontal="center")
    for idx, term in enumerate(quote_defaults["payment_terms"], start=1):
        ws[f"C{payment_start + idx}"] = term
        ws[f"C{payment_start + idx}"].alignment = Alignment(wrap_text=True)

    ws.sheet_view.showGridLines = False

    detail_sheet = workbook.create_sheet("Detalle BOM")
    detail_sheet.append(["Category", "Item", "Unit", "Quantity", "Unit Price", "Line Total", "Source"])
    for line in quote["bom"]:
        detail_sheet.append(
            [line["category"], line["item"], line["unit"], line["quantity"], line["unit_price"], line["line_total"], line["source"]]
        )
    detail_sheet.sheet_state = "hidden"

    assumptions_sheet = workbook.create_sheet("Supuestos")
    assumptions_sheet.append(["Assumption"])
    for item in quote["assumptions"]:
        assumptions_sheet.append([item])
    assumptions_sheet.sheet_state = "hidden"

    live_sheet = workbook.create_sheet("Precios Vivos")
    live_sheet.append(["Key", "Label", "Price", "Fetched At", "Fallback Used", "Source URL"])
    for key, item in quote["live_prices"].items():
        live_sheet.append([key, item["label"], item["price"], item["fetched_at"], item["fallback_used"], item["source_url"]])
    live_sheet.sheet_state = "hidden"

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_pdf(quote: dict[str, Any], plan_path: str | Path, output_path: str | Path, company: dict[str, Any]) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    c = canvas.Canvas(str(path), pagesize=A4)
    _draw_pdf_content(c, quote, company, plan_image_path=str(plan_path))
    c.save()
    return path


def export_pdf_bytes(quote: dict[str, Any], plan_png_bytes: bytes, company: dict[str, Any]) -> bytes:
    """Generate PDF in memory and return raw bytes (for serverless / streaming)."""
    from io import BytesIO
    from reportlab.lib.utils import ImageReader

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    # Write plan PNG to a temporary ImageReader
    plan_reader = ImageReader(BytesIO(plan_png_bytes)) if plan_png_bytes else None
    _draw_pdf_content(c, quote, company, plan_image_reader=plan_reader)
    c.save()
    return buf.getvalue()


def _draw_pdf_content(
    c,
    quote: dict[str, Any],
    company: dict[str, Any],
    plan_image_path: str | None = None,
    plan_image_reader=None,
) -> None:
    """Shared PDF drawing logic for both file and in-memory modes."""
    width, height = A4
    margin = 18 * mm

    c.setFillColor(colors.HexColor("#243B2A"))
    c.rect(0, height - 38 * mm, width, 38 * mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(margin, height - 18 * mm, company["name"])
    c.setFont("Helvetica", 11)
    c.drawString(margin, height - 25 * mm, company["tagline"])

    project = quote["project"]
    geometry = quote["geometry"]
    totals = quote["totals"]

    y = height - 50 * mm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, f"Cotizacion {project.get('quote_id', 'MVP')}")
    y -= 8 * mm

    c.setFont("Helvetica", 11)
    lines = [
        f"Proyecto: {project['project_name']}",
        f"Cliente: {project['client']['name']}",
        f"Ubicacion: {project['location']}",
        f"Fecha: {quote['generated_on']}",
        f"Superficie estimada: {geometry.gross_area_m2:.2f} m2",
        f"Precio final estimado: {format_clp(totals['final_total'])}",
        f"Valor por m2: {format_clp(totals['price_per_m2'])}",
    ]
    for line in lines:
        c.drawString(margin, y, line)
        y -= 6 * mm

    # Draw plan image (from file path or in-memory reader)
    img_source = plan_image_path or plan_image_reader
    if img_source:
        c.drawImage(img_source, margin, y - 85 * mm, width=95 * mm, height=95 * mm, preserveAspectRatio=True, mask="auto")

    breakdown_x = 122 * mm
    breakdown_y = height - 77 * mm
    c.setFont("Helvetica-Bold", 13)
    c.drawString(breakdown_x, breakdown_y, "Resumen economico")
    breakdown_y -= 8 * mm
    c.setFont("Helvetica", 10)
    breakdown = [
        ("Materiales", totals["materials_subtotal"]),
        ("Fabricacion", totals["manufacturing_cost"]),
        ("Instalacion", totals["installation_cost"]),
        ("Fundaciones", totals["foundation_cost"]),
        ("Transporte", totals["transport_cost"]),
        ("Ingenieria", totals["engineering_cost"]),
        ("Puesta en obra", totals["site_setup_cost"]),
        ("Contingencia", totals["contingency"]),
        ("Overhead", totals["overhead"]),
        ("Margen", totals["profit"]),
    ]
    for label, value in breakdown:
        c.drawString(breakdown_x, breakdown_y, label)
        c.drawRightString(width - margin, breakdown_y, format_clp(value))
        breakdown_y -= 5.5 * mm

    notes_y = 62 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, notes_y, "Supuestos clave")
    notes_y -= 7 * mm
    c.setFont("Helvetica", 10)
    for item in quote["assumptions"]:
        c.drawString(margin, notes_y, f"- {item}")
        notes_y -= 5 * mm

    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#4D4D4D"))
    c.drawString(margin, 12 * mm, "MVP generado por Modul CAD. Valores referenciales sujetos a validacion comercial.")

